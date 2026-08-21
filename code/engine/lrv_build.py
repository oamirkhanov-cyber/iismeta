# -*- coding: utf-8 -*-
"""IISMETA cost-estimation engine module."""
import io
import re
import collections


def _num(s, default=0.0):
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s or "").strip().replace(",", ".")
    if not s:
        return default
    try:
        return float(s)
    except ValueError:
        m = re.search(r"-?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", s)
        return float(m.group(0)) if m else default


def _ru(x):
    """IISMETA cost-estimation engine module."""
    if x is None or x == "":
        return ""
    try:
        f = float(str(x).replace(",", "."))
    except ValueError:
        return str(x)
    s = ("%.8f" % f).rstrip("0").rstrip(".")
    return (s or "0").replace(".", ",")


def _num(x):
    """IISMETA cost-estimation engine module."""
    if x is None or x == "":
        return None
    try:
        f = float(str(x).replace(",", "."))
    except (ValueError, TypeError):
        return str(x)
    return round(f, 8)


def _uu(s):
    """IISMETA cost-estimation engine module."""
    s = (s or "").upper().replace("³", "3").replace("²", "2")
    return re.sub(r"^(\d{2,}) ([А-Я]\d?)$", r"\1\2", s)


def _izm_mult(izm):
    """IISMETA cost-estimation engine module."""
    m = re.match(r"\s*(\d+(?:[.,]\d+)?)", str(izm or ""))
    return float(m.group(1).replace(",", ".")) if m else 1.0


def _vol_suspect(izm, norm_qty):
    """IISMETA cost-estimation engine module."""
    try:
        q = float(norm_qty)
    except (TypeError, ValueError):
        return None
    if q <= 0:
        return None
    u = str(izm or "").lower()
    if re.search(r"\bт\b|тонн|т конструкц", u) and not re.search(r"м2|м3|шт", u):
        cap = 20.0
    elif re.match(r"\s*1000", u):
        cap = 100.0
    elif re.match(r"\s*100", u):
        cap = 80.0
    else:
        cap = 300.0
    if q > cap:
        return "⚠ ОБЪЁМ %s измерителей аномально велик — проверьте (возможно ×N/×100/спутана ед.изм)" % (
            round(q, 1))
    return None


def _is_labor(unit):
    u = (unit or "").lower()
    return "чел" in u or "маш" in u


def res_type(шифр, unit):
    """IISMETA cost-estimation engine module."""
    u = (unit or "").lower()
    if "чел" in u:
        return "труд"
    if "маш" in u:
        return "машины"
    if u.strip():
        return "материалы"
    s = str(шифр or "").strip().upper()
    if re.fullmatch(r"\d+", s):
        return "труд"
    if s.startswith("М"):
        return "материалы"
    if s.startswith("С"):
        return "машины"
    return "материалы"


RES_COLOR = {"труд": "7A4FA0", "машины": "1F8A5B", "материалы": "1F3864"}


def _strip_code(code):
    """IISMETA cost-estimation engine module."""
    s = str(code or "").strip()
    m = re.sub(r"^[^\d]+", "", s)
    return m or s


def _coef_inline(coefs):
    """IISMETA cost-estimation engine module."""
    if not coefs:
        return ""
    parts = []
    for c in coefs:
        seg = "п.%s" % c.get("пункт", "")
        kt = str(c.get("кзт", c.get("коэф", ""))).strip()
        km = str(c.get("кэм", "")).strip()
        kk = []
        if kt and kt not in ("-", ""):
            kk.append("Кзт=%s" % kt)
        if km and km not in ("-", ""):
            kk.append("Кэм=%s" % km)
        parts.append(seg + ((" " + " ".join(kk)) if kk else ""))
    return " · попр. " + "; ".join(parts)


def _popr_text(coefs):
    """IISMETA cost-estimation engine module."""
    if not coefs:
        return ""
    out = []
    for c in coefs:
        usl = str(c.get("усл", "") or "").strip().rstrip(".")
        kt = str(c.get("кзт", c.get("коэф", "")) or "").strip()
        km = str(c.get("кэм", "") or "").strip()
        ph = []
        if kt and kt not in ("-", ""):
            ph.append("к нормам затрат труда-%s" % kt)
        if km and km not in ("-", ""):
            ph.append("нормам эксплуатации машин-%s" % km)
        seg = usl
        if ph:
            seg = ((usl + ", ") if usl else "") + "применён коэффициент " + " и ".join(ph)
        if seg:
            out.append(seg)
    if not out:
        return ""
    sfx = "; ".join(out)
    sfx = sfx[:1].upper() + sfx[1:]
    return ". " + sfx


def fill_unit(шифр, unit):
    """IISMETA cost-estimation engine module."""
    if unit and unit.strip():
        return unit
    return {"труд": "чел.-ч", "машины": "маш.-ч"}.get(res_type(шифр, unit), unit or "")


def _clean(s):
    """IISMETA cost-estimation engine module."""
    s = re.sub(r"-\s+", "", s or "")
    return re.sub(r"\s+", " ", s).strip()


def _sostav_items(s):
    """IISMETA cost-estimation engine module."""
    s = (s or "").strip()
    if not s:
        return []
    parts = re.split(r"\s+(?=\d{1,2}\s*\.)", s)
    return [p.strip() for p in parts if p.strip()]


def build_lrv(resolved):
    """IISMETA cost-estimation engine module."""
    works = []
    for res in resolved:
        if res.get("sec"):
            works.append({"sec": True, "kind": res.get("kind", ""), "наим": res.get("name", ""),
                          "num": res.get("num", ""),
                          "код": "", "попр": "", "измеритель": "", "объём": "", "кол_изм": "",
                          "Кр": "", "статус": "", "коэфф": [], "ресурсы": [], "состав": "",
                          "вариант": "", "доп": False, "примечание": "", "раздел": ""})
            continue
        if res.get("cond"):
            works.append({"cond": True, "kzt": _num(res.get("kzt"), 1.0), "kem": _num(res.get("kem"), 1.0),
                          "num": res.get("num", ""), "usl": res.get("usl", ""), "text": res.get("text", ""),
                          "section": res.get("section", ""), "group": res.get("group", "MAIN"),
                          "код": "", "попр": "", "измеритель": "",
                          "объём": "", "кол_изм": "", "Кр": "", "статус": "", "коэфф": [],
                          "ресурсы": [], "состав": "", "вариант": "", "доп": False, "раздел": ""})
            continue
        if res.get("err"):
            works.append({"err": True, "наим": res.get("name", ""), "wc": res.get("wc", ""),
                          "код": res.get("code", "") or "—", "попр": "", "измеритель": "",
                          "объём": "", "кол_изм": "", "Кр": "", "статус": "", "коэфф": [],
                          "ресурсы": [], "состав": "", "вариант": "", "доп": False, "примечание": "", "раздел": ""})
            continue
        code = res["вариант"]["code"] or res["вид"]["work_code"]
        if not code:
            continue
        izm = res["измеритель"]
        mult = _izm_mult(izm)
        pv0 = res.get("перевозка_расценка")
        if pv0 and pv0.get("код") and pv0.get("объём_тонн") is not None:
            qty = pv0["объём_тонн"]
        else:
            qty = _num(res.get("объём"))
        norm_qty = (qty / mult) if mult else qty
        coefs = res.get("коэффициенты", [])
        if res.get("_baked"):
            Kr = 1.0
        else:
            Kr = 1.0
            for c in coefs:
                Kr *= _num(c.get("коэф"), 1.0) or 1.0
        rows = []
        for r in res.get("ресурсы", []):
            val = _num(r.get("val"))
            if val <= 0:
                continue
            sh = r.get("res", "")
            unit = fill_unit(sh, r.get("unit", ""))
            k = Kr if _is_labor(unit) else 1.0
            kolvo = round(val * norm_qty * k, 6)
            row = {"шифр": sh, "наим": _clean(r.get("name", "")),
                   "ед": unit, "норма": val, "кол": kolvo}
            if r.get("ppro"):
                row["по_проекту"] = True
            rows.append(row)
        pv = res.get("перевозка_расценка")
        nm_work = (pv["наим"] if (pv and pv.get("код")) else (res.get("вход") or res["вид"]["имя"]))
        popr_txt = _popr_text(coefs)
        _vs = _vol_suspect(izm, norm_qty)
        _dq, _dk = res.get("_disp_objem"), res.get("_disp_k")
        if _dq is not None and _dk:
            qty = _dq
            norm_qty = _num(_dq) / (_izm_mult(izm) or 1.0)
            Kr = _dk
        works.append({"код": code, "наим": nm_work, "попр": popr_txt,
                      "измеритель": izm, "объём": qty, "кол_изм": round(norm_qty, 4),
                      "Кр": round(Kr, 4), "статус": ("🔴" if _vs else res["статус"]), "коэфф": coefs, "ресурсы": rows,
                      "состав": res.get("состав", "") or "", "вариант": res.get("вариант_доп", "") or "",
                      "доп": False, "примечание": (_vs or ""), "раздел": res.get("раздел") or "Земляные работы"})
        for dp in res.get("доп_позиции", []):
            if dp.get("_dobor"):
                n = dp.get("n", 0)
                d_rows = []
                for r in dp.get("ресурсы", []):
                    val = _num(r.get("val"))
                    if val <= 0:
                        continue
                    unit = fill_unit(r.get("res", ""), r.get("unit", ""))
                    d_rows.append({"шифр": r.get("res", ""), "наим": _clean(r.get("name", "")),
                                   "ед": unit, "норма": val, "кол": round(val * norm_qty * n, 6)})
                works.append({"код": dp.get("код") or "—", "наим": dp.get("наим", ""),
                              "измеритель": izm, "объём": qty, "кол_изм": round(norm_qty, 4),
                              "Кр": n, "статус": "🟡", "коэфф": [], "ресурсы": d_rows,
                              "доп": True, "примечание": dp.get("примечание", ""),
                              "раздел": res.get("раздел") or "Земляные работы"})
                continue
            тонн = dp.get("объём_тонн")
            dp_rows = []
            for r in dp.get("resources", []):
                v = _num(r.get("val"))
                if v <= 0:
                    continue
                kol = round(v * тонн, 6) if тонн else v
                dp_rows.append({"шифр": r.get("res", ""), "наим": _clean(r.get("name", "")),
                                "ед": r.get("unit", ""), "норма": r.get("val", ""), "кол": kol})
            works.append({
                "код": dp.get("код") or "—", "наим": dp["наим"],
                "измеритель": "1 т" if тонн else dp.get("ед", ""),
                "объём": тонн if тонн else dp.get("L_проект", ""), "кол_изм": тонн or "",
                "Кр": dp.get("коэффициент") or 1.0, "статус": dp.get("статус", "🟡"),
                "коэфф": [], "ресурсы": dp_rows, "доп": True, "примечание": dp.get("примечание", ""),
                "раздел": res.get("раздел") or "Земляные работы"})
    return works


def apply_usloviya(works):
    """IISMETA cost-estimation engine module."""
    cond_targets = []
    for i, cw in enumerate(works):
        if not cw.get("cond"):
            continue
        kzt = _num(cw.get("kzt"), 1.0) or 1.0
        kem = _num(cw.get("kem"), 1.0) or 1.0
        k = i - 1
        while k >= 0 and works[k].get("cond"):
            k -= 1
        anchor = works[k] if k >= 0 else None
        scoped = bool(anchor and anchor.get("sec"))
        cw["_scope"] = "раздел" if scoped else "вся смета"
        if scoped:
            anchor_sub = (anchor.get("kind") == "Подраздел")
            targets = []
            for j in range(i + 1, len(works)):
                w = works[j]
                if w.get("sec") and (anchor_sub or w.get("kind") != "Подраздел"):
                    break
                if w.get("sec") or w.get("cond"):
                    continue
                targets.append(w)
        else:
            targets = [w for w in works if not w.get("sec") and not w.get("cond")]
        cond_targets.append((cw, str(cw.get("group", "MAIN")).upper(),
                             frozenset(id(w) for w in targets)))
        if kzt == 1.0 and kem == 1.0:
            continue
        for w in targets:
            for r in w.get("ресурсы", []):
                tp = res_type(str(r.get("шифр", "")).strip(), r.get("ед", ""))
                if tp == "труд" and kzt != 1.0:
                    r["норма"] = round(_num(r.get("норма")) * kzt, 6)
                    r["кол"] = round(_num(r.get("кол")) * kzt, 6)
                elif tp == "машины" and kem != 1.0:
                    r["норма"] = round(_num(r.get("норма")) * kem, 6)
                    r["кол"] = round(_num(r.get("кол")) * kem, 6)
    for a, (cwa, ga, ta) in enumerate(cond_targets):
        if ga != "MAIN" or not ta:
            continue
        for b, (cwb, gb, tb) in enumerate(cond_targets):
            if b != a and gb == "MAIN" and (ta & tb):
                cwa["_warn"] = "не совмещается с другим основным условием — по нормам применяется одно"
                break
    return works


def lrv_totals(works):
    """IISMETA cost-estimation engine module."""
    votes = {}
    for w in works:
        for r in w["ресурсы"]:
            if r["ед"]:
                votes.setdefault(str(r["шифр"]).strip(), collections.Counter())[r["ед"]] += 1
    canon = {sh: c.most_common(1)[0][0] for sh, c in votes.items()}
    agg = {}
    for w in works:
        for r in w["ресурсы"]:
            sh = str(r["шифр"]).strip()
            key = (sh, r["ед"] or canon.get(sh, ""))
            a = agg.setdefault(key, {"кол": 0.0, "наим": ""})
            a["кол"] = round(a["кол"] + r["кол"], 6)
            if len(r["наим"]) > len(a["наим"]):
                a["наим"] = r["наим"]
    out = [{"шифр": k[0], "наим": v["наим"], "ед": k[1], "кол": v["кол"]} for k, v in agg.items()]
    out.sort(key=lambda x: x["шифр"])
    return out


def lrv_summary(works):
    """IISMETA cost-estimation engine module."""
    labor = 0.0
    machine = 0.0
    mats = {}
    for w in works:
        for r in w["ресурсы"]:
            tp = res_type(r["шифр"], r["ед"])
            if tp == "труд":
                labor = round(labor + r["кол"], 6)
            elif tp == "машины":
                machine = round(machine + r["кол"], 6)
            else:
                key = (str(r["шифр"]).strip(), r["ед"])
                a = mats.setdefault(key, {"кол": 0.0, "наим": ""})
                a["кол"] = round(a["кол"] + r["кол"], 6)
                if len(r["наим"]) > len(a["наим"]):
                    a["наим"] = r["наим"]
    out = []
    if labor:
        out.append({"раздел": "Труд", "шифр": "", "наим": "Затраты труда рабочих и машинистов, всего",
                    "ед": "чел.-ч", "кол": labor})
    if machine:
        out.append({"раздел": "Машины", "шифр": "", "наим": "Эксплуатация машин и механизмов, всего",
                    "ед": "маш.-ч", "кол": machine})
    for (sh, ed), a in sorted(mats.items()):
        out.append({"раздел": "Материалы", "шифр": sh, "наим": a["наим"], "ед": ed, "кол": a["кол"]})
    return out


LABOR_NAME = {"000001": "Затраты труда рабочих-строителей", "000003": "Затраты труда машинистов"}


def res_vedomost(works):
    """IISMETA cost-estimation engine module."""
    labor, mach, mat = {}, {}, {}
    for w in works:
        for r in w["ресурсы"]:
            sh = str(r["шифр"]).strip()
            tp = res_type(sh, r["ед"])
            if tp == "труд":
                nm = (r["наим"] or "").lower()
                if "машинист" in nm:
                    code = "000003"
                elif "рабоч" in nm:
                    code = "000001"
                else:
                    code = "000001" if sh in ("1", "01") else ("000003" if sh in ("3", "03") else sh)
                a = labor.setdefault(code, {"наим": LABOR_NAME.get(code, _clean(r["наим"])), "ед": "чел.-ч", "кол": 0.0})
                a["кол"] = round(a["кол"] + r["кол"], 6)
            else:
                tgt = mach if tp == "машины" else mat
                ed = fill_unit(sh, r["ед"]) if tp == "машины" else r["ед"]
                key = (sh, ed)
                a = tgt.setdefault(key, {"наим": "", "кол": 0.0})
                a["кол"] = round(a["кол"] + r["кол"], 6)
                if len(r["наим"]) > len(a["наим"]):
                    a["наим"] = r["наим"]
    def numk(c):
        m = re.match(r"\d+", str(c)); return int(m.group(0)) if m else 0
    def rows(d, keyed):
        if keyed:
            return [{"шифр": k[0], "наим": v["наим"], "ед": k[1], "кол": v["кол"]}
                    for k, v in sorted(d.items(), key=lambda kv: (numk(kv[0][0]), kv[0][1]))]
        return [{"шифр": c, "наим": v["наим"], "ед": v["ед"], "кол": v["кол"]}
                for c, v in sorted(d.items(), key=lambda kv: numk(kv[0]))]
    return [("ТРУДОВЫЕ РЕСУРСЫ", rows(labor, False)),
            ("СТРОИТЕЛЬНЫЕ МАШИНЫ И МЕХАНИЗМЫ", rows(mach, True)),
            ("МАТЕРИАЛЬНЫЕ РЕСУРСЫ", rows(mat, True))]


def _forma5_top(ws, glob, ncols, title):
    """IISMETA cost-estimation engine module."""
    from openpyxl.styles import Font, Alignment, Border, Side
    hair = Side(style="hair", color="000000")
    under = Border(bottom=hair)
    стройка = glob.get("стройка") or ""
    объект = glob.get("объект") or ""
    основание = glob.get("основание") or ""
    номер = glob.get("номер") or ""
    f9 = Font(name="Times New Roman", size=9)
    f8 = Font(name="Times New Roman", size=8)
    f12 = Font(name="Times New Roman", size=12, bold=True)

    def line(r, text, font, align="center", border=None):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(r, 1, text); c.font = font; c.alignment = Alignment(align, "center", wrap_text=True)
        if border:
            for j in range(1, ncols + 1):
                ws.cell(r, j).border = border

    fc = ws.cell(1, ncols, "Форма N 5"); fc.font = f8; fc.alignment = Alignment("right", "center")
    line(2, стройка, f9, "center", under); line(3, "(наименование стройки)", f8, "center")
    line(4, ("%s   № %s" % (title, номер)).rstrip(), f12); line(5, "(локальная ресурсная смета)", f8)
    line(6, "на   %s" % объект, f9, "center", under)
    line(7, "(наименование работ и затрат, наименование объекта)", f8)
    line(8, "Основание:   %s" % основание, f9, "left", under)
    return 9


def to_excel_res(works, glob, sbornik="ШНК 4.02.01-04 «Земляные работы»"):
    """IISMETA cost-estimation engine module."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    NAVY, BLUE = "1F3864", "2E75B6"
    wb = Workbook(); ws = wb.active; ws.title = "RES"
    thin = Side(style="thin", color="C8D4E6"); box = Border(thin, thin, thin, thin)
    hdr = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    navy = PatternFill("solid", fgColor=NAVY); lite = PatternFill("solid", fgColor="D9E2F3")
    ctr = Alignment("center", "center", wrap_text=True)

    r0 = _forma5_top(ws, glob, 7, "ЛОКАЛЬНЫЙ РЕСУРСНЫЙ СМЕТНЫЙ РАСЧЕТ")
    titles = ["№ п.п.", "Шифр номера нормативов и коды ресурсов", "Наименование работ и затрат",
              "Единица измерения", "Количество"]
    widths = [6, 18, 44, 12, 12, 14, 14]
    for j, (h, w) in enumerate(zip(titles, widths), 1):
        ws.merge_cells(start_row=r0, start_column=j, end_row=r0 + 1, end_column=j)
        c = ws.cell(r0, j, h); c.font = hdr; c.fill = navy; c.alignment = ctr; c.border = box
        ws.column_dimensions[chr(64 + j)].width = w
    ws.merge_cells(start_row=r0, start_column=6, end_row=r0, end_column=7)
    c = ws.cell(r0, 6, "Сметная стоимость в базисном уровне"); c.font = hdr; c.fill = navy; c.alignment = ctr; c.border = box
    for j, sub in ((6, "на ед. изм."), (7, "общая")):
        c = ws.cell(r0 + 1, j, sub); c.font = hdr; c.fill = navy; c.alignment = ctr; c.border = box
        ws.column_dimensions[chr(64 + j)].width = widths[j - 1]
    ws.row_dimensions[r0].height = 28
    for j in range(1, 8):
        c = ws.cell(r0 + 2, j, j); c.font = Font(name="Arial", bold=True, color=NAVY, size=9)
        c.fill = lite; c.alignment = Alignment("center", "center"); c.border = box

    row = r0 + 3
    for title, items in res_vedomost(works):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row, 1, title)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10); c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment("left", "center"); c.border = box
        row += 1
        for n, it in enumerate(items, 1):
            vals = [n, it["шифр"], it["наим"], it["ед"], it["кол"], "", ""]
            for j, v in enumerate(vals, 1):
                cc = ws.cell(row, j, v); cc.font = Font(name="Arial", size=9); cc.border = box
                cc.alignment = Alignment("center" if j != 3 else "left", "center", wrap_text=(j in (3, 4)))
            row += 1
    ws.freeze_panes = "A%d" % (r0 + 3)
    _autofit_rows(ws, {3: widths[2]})
    _setup_a4(ws, 7)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def _autofit_rows(ws, wrapcols, base=15.0):
    """IISMETA cost-estimation engine module."""
    import math
    for r in range(1, ws.max_row + 1):
        lines = 1
        for col, width in wrapcols.items():
            v = ws.cell(r, col).value
            if v is not None and not isinstance(v, (int, float)):
                v = v if isinstance(v, str) else str(v)
                for seg in v.split("\n"):
                    lines = max(lines, max(1, math.ceil(len(seg) / max(8, width * 0.95))))
        if lines > 1:
            ws.row_dimensions[r].height = round(lines * base, 1)


def _setup_a4(ws, ncols=6):
    """IISMETA cost-estimation engine module."""
    from openpyxl.worksheet.properties import PageSetupProperties
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9                # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)
    ws.print_area = "A1:%s%d" % (chr(64 + ncols), ws.max_row)


def to_excel(works, glob, sbornik="", show_sostav=False):
    """IISMETA cost-estimation engine module."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    TNR = "Times New Roman"
    hair = Side(style="hair", color="000000")
    box = Border(left=hair, right=hair, top=hair, bottom=hair)
    cyan = PatternFill("solid", fgColor="CCFFFF")
    f9 = Font(name=TNR, size=9)
    f9b = Font(name=TNR, size=9, bold=True)
    f10b = Font(name=TNR, size=10, bold=True)
    f10 = Font(name=TNR, size=10)
    wb = Workbook(); ws = wb.active; ws.title = "ЛРВ"

    def AL(h="left", wrap=False, v="top"):
        return Alignment(h, v, wrap_text=wrap)

    def put(r, c, v, font=f9, al=None, border=True):
        cell = ws.cell(r, c, v); cell.font = font
        cell.alignment = al or AL("left")
        if border:
            cell.border = box
        return cell

    r0 = _forma5_top(ws, glob, 6, "ЛОКАЛЬНАЯ РЕСУРСНАЯ ВЕДОМОСТЬ")
    for col, px in zip("ABCDEF", [38, 95, 580, 71, 56, 56]):
        ws.column_dimensions[col].width = round(px / 7.0, 1)
    for j, h in enumerate(["N п.п.", "Шифр номера нормативов и коды ресурсов",
                           "Наименование работ и затрат", "Единица измерения"], 1):
        ws.merge_cells(start_row=r0, start_column=j, end_row=r0 + 1, end_column=j)
        put(r0, j, h, f9, AL("center", True))
        ws.cell(r0 + 1, j).border = box
    ws.merge_cells(start_row=r0, start_column=5, end_row=r0, end_column=6)
    put(r0, 5, "Количество", f9, AL("center", True)); ws.cell(r0, 6).border = box
    put(r0 + 1, 5, "на. ед. измерения", f9, AL("center", True))
    put(r0 + 1, 6, "по проектным данным", f9, AL("center", True))
    for j in range(1, 7):
        put(r0 + 2, j, j, f9b, AL("center"))
    for rr in (r0, r0 + 1, r0 + 2):
        for j in range(1, 7):
            cl = ws.cell(rr, j); cl.fill = cyan
            cl.alignment = Alignment(cl.alignment.horizontal, "center", wrap_text=cl.alignment.wrap_text)

    row = r0 + 3
    n = 0
    for w in works:
        if w.get("sec"):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            kind = w.get("kind", ""); nm = (w.get("наим", "") or "").upper()
            num = w.get("num", "")
            head = ((kind + (" " + num if num else "") + ": " + nm) if kind else nm)
            put(row, 1, head, f10b, AL("left"))
            for j in range(1, 7):
                ws.cell(row, j).border = box; ws.cell(row, j).fill = cyan
            row += 1
            continue
        if w.get("cond"):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ktxt = "Кзтр %s" % _ru(w.get("kzt"))
            if _num(w.get("kem"), 1.0) != 1.0:
                ktxt += " · Кэм %s" % _ru(w.get("kem"))
            body = (w.get("text") or w.get("usl") or "")
            txt = "УСЛОВИЯ ПРОИЗВОДСТВА РАБОТ (%s · %s): %s" % (ktxt, w.get("_scope", "вся смета"), body)
            put(row, 1, txt.upper(), f9b, AL("left", True))
            for j in range(1, 7):
                ws.cell(row, j).border = box; ws.cell(row, j).fill = cyan
            row += 1
            continue
        if w.get("err"):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            put(row, 1, ("!! РАСЦЕНКА НЕ НАЙДЕНА В БАЗЕ (подберите заново): %s" % (w.get("наим", "") or "")).upper(),
                f9b, AL("left", True))
            for j in range(1, 7):
                ws.cell(row, j).border = box
            row += 1
            continue
        n += 1
        code = str(w["код"]) or "—"
        is_extra = code.upper() == "ВНЕ"
        if is_extra:
            res0 = (w["ресурсы"] or [{}])[0]
            shifr = str(res0.get("шифр") or "—").strip()
        else:
            shifr = ("Е" + code) if re.match(r"\d", code) else code
            if w.get("вариант"):
                shifr += " " + w["вариант"]
            try:
                _kr = float(str(w.get("Кр") or 1).replace(",", "."))
            except (TypeError, ValueError):
                _kr = 1.0
            if abs(_kr - 1.0) > 1e-9:
                shifr += "К=%s" % _ru(_kr)
        nm_cell = ((w["наим"] or "") + (w.get("попр", "") or "")).upper()
        if w.get("примечание"):
            nm_cell += "   " + w["примечание"]
        put(row, 1, n, f10b, AL("center")); put(row, 2, shifr, f10b, AL("left"))
        put(row, 3, nm_cell, f10b, AL("left", True))
        put(row, 4, _uu(w["измеритель"]), f10b, AL("center"))
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        put(row, 5, _num(w.get("кол_изм")), f10b, AL("center")); ws.cell(row, 6).border = box
        row += 1
        if is_extra:
            continue
        items = _sostav_items(w.get("состав", "")) if show_sostav else []
        if items:
            for j in (1, 2, 4, 5, 6):
                put(row, j, None)
            put(row, 3, "СОСТАВ РАБОТ:", f9, AL("left"))
            row += 1
            for it in items:
                for j in (1, 2, 4, 5, 6):
                    put(row, j, None)
                put(row, 3, it, f9, AL("left", True))
                row += 1
        for k, r in enumerate(w["ресурсы"], 1):
            put(row, 1, "%d.%d" % (n, k), f9, AL("center"))
            put(row, 2, _strip_code(r["шифр"]), f9, AL("center"))
            put(row, 3, (r["наим"] or "").upper(), f9, AL("left", True))
            put(row, 4, _uu(r["ед"]), f9, AL("center"))
            put(row, 5, _num(r["норма"]), f9, AL("right"))
            put(row, 6, _num(r["кол"]), f9, AL("right"))
            row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    put(row, 1, "ИТОГО ПО ЛОКАЛЬНОЙ РЕСУРСНОЙ ВЕДОМОСТИ:", f9b, AL("left"))
    for j in range(1, 7):
        ws.cell(row, j).border = box; ws.cell(row, j).fill = cyan
    row += 1
    jj = 0
    for title, items in res_vedomost(works):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        put(row, 1, title, f9b, AL("left"))
        for j in range(1, 7):
            ws.cell(row, j).border = box; ws.cell(row, j).fill = cyan
        row += 1
        for it in items:
            jj += 1
            sc = _strip_code(it["шифр"]); sc = sc.lstrip("0") or sc
            put(row, 1, jj, f9, AL("center"))
            put(row, 2, sc, f9, AL("left"))
            put(row, 3, (it["наим"] or "").upper(), f9, AL("left", True))
            put(row, 4, _uu(it["ед"]), f9, AL("center"))
            put(row, 5, None)
            put(row, 6, _num(it["кол"]), f10, AL("right"))
            row += 1

    ws.freeze_panes = "A%d" % (r0 + 3)
    _autofit_rows(ws, {3: round(580 / 7.0, 1)})
    _setup_a4(ws, 6)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
